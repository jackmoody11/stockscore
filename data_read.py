# Modules
import openpyxl as xl
import pandas as pd
import datetime
import requests

# Variables
iex_url_base = "https://api.iextrading.com/1.0/"


# Functions

## Symbols
def xl_get_symbols():

    file = 'stock_score_data.xlsx'
    wb = xl.load_workbook(file)
    ws = wb['Data']
    symbols = []

    for i in range(4,6000):
        if ws['A%s'%i].value != None:
            symbols.append(ws['A%s'%i].value)
    return symbols


def get_symbols(iex_url_base = iex_url_base, manual = False):

    """Gets all symbols from IEX API (uppercase) or Excel
    if saved within last day. """

    file = 'stock_score_data.xlsx'
    wb = xl.load_workbook(file)
    ws = wb['Data']

    last_updated = ws['B1'].value
    time_now = datetime.datetime.now()
    time_diff = time_now - last_updated

    if (last_updated == None or time_diff > datetime.timedelta(days=5) or manual):

        symbols_json = requests.get(iex_url_base + "ref-data/symbols").json()
        symbols = []
        for i in range(len(symbols_json)):
            if symbols_json[i]['type'] == 'cs':
                symbols.append(symbols_json[i]['symbol'])

    else:

        symbols = xl_get_symbols()

    return symbols


def xl_update_symbols():

    file = 'stock_score_data.xlsx'
    wb = xl.load_workbook(file)
    ws = wb['Data']

    last_updated = ws['B1'].value
    time_now = datetime.datetime.now()
    time_diff = time_now - last_updated

    if (last_updated == None or time_diff > datetime.timedelta(days=5)):
        ws['B1'].value = datetime.datetime.now()
        last_updated = datetime.datetime.now()
        symbols = get_symbols(manual = True)
        for i in range(4,len(symbols)+4):
            ws['A%s'%i] = symbols[i - 4]

    wb.save(file)
    # print("The new last updated time is " + str(last_updated))
    # print(ws['A10'].value)


##
def set_batches(symbols):

    """Calculate number of batches to send for GET request.

    Creates batches of 100 tickers to limit number of GET requests sent to IEX.
    """
    num_batches = int(len(symbols) / 100 + round(len(symbols) % 100))
    x = 0
    batch_symbols = {}
    for i in range(0, num_batches):
        if(x + 101 < len(symbols)):
            batch_symbols[i] = ",".join(symbols[x:x + 100])
        else:
            batch_symbols[i] = ",".join(symbols[x:len(symbols) + 1])
            break
        x = (i + 1) * 100 + 1

    return batch_symbols

def xl_update_stats(manual = False):

    file = 'stock_score_data.xlsx'
    wb = xl.load_workbook(file)
    ws = wb['Data']

    last_updated = ws['B1'].value
    time_now = datetime.datetime.now()
    time_diff = time_now - last_updated

    if (last_updated == None or time_diff > datetime.timedelta(days=1) or manual):
        ws['B1'].value = datetime.datetime.now()
        last_updated = datetime.datetime.now()
        symbols = get_symbols()
        batch_symbols = set_batches(symbols)
        result = get_stats(batch_symbols)

        for symbol in result:
            if(result[symbol.upper()].get('stats')):
                base = result[symbol]['stats']
                # Price to book test -- Want a lower price to book
                if(base['priceToBook']):
                    score = round(5 / (base['priceToBook'] + 0.8))
                    if(0 < base['priceToBook'] <= 1):
                        stock_scores[symbol] += score
                        print(symbol + " score went up by " + str(score) + "-- price to book between 0 and 1")
                    elif(1 < base['priceToBook'] <= 2):
                        stock_scores[symbol] += score
                        print(symbol + " score went up by " + str(score) + "-- price to book between 1 and 2")


