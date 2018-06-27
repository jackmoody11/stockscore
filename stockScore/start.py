# Modules
from bs4 import BeautifulSoup
import requests
import openpyxl as xl
import datetime

iex_url_base = "https://api.iextrading.com/1.0/"

# Functions
def init_stock_scores(symbols):

	"""Set all stock scores to zero. """
	stock_scores = {}
	for symbol in symbols:
	    stock_scores[symbol] = 0
	return stock_scores


def xl_get_symbols():
	"""Assumes symbols are already in Excel. """
    file = 'stock_score_data.xlsx'
    wb = xl.load_workbook(file)
    ws = wb['Data']
    symbols = []

    for i in range(4,6000):
        if ws['A%s'%i].value != None:
            symbols.append(ws['A%s'%i].value)
    return symbols


def get_symbols(iex_url_base = iex_url_base, manual = False):

    """Gets all symbols from IEX API (uppercase) or Excel
    if saved within last five days. To manually force refresh,
    set manual = True. """

    file = 'stock_score_data.xlsx'
    wb = xl.load_workbook(file)
    ws = wb['Data']

    last_updated = ws['B1'].value
    time_now = datetime.datetime.now()
    time_diff = time_now - last_updated

    if (last_updated == None or time_diff > datetime.timedelta(days=5) or manual):

        symbols_json = requests.get(iex_url_base + "ref-data/symbols").json()
        symbols = []
        for i in range(len(symbols_json)):
            if symbols_json[i]['type'] == 'cs':
                symbols.append(symbols_json[i]['symbol'])
                for i in range(4,len(symbols)+4):
            		ws['A%s'%i] = symbols[i - 4]
            	wb.save(file)

    else:

        symbols = xl_get_symbols()

    return symbols


def set_batches(symbols):

	"""Calculate number of batches to send for GET request.
	Creates batches of 100 symbols to limit number of GET requests sent to IEX.
	"""
	num_batches = int(len(symbols) / 100 + round(len(symbols) % 100))
	x = 0
	batch_symbols = {}
	for i in range(0, num_batches):
	    if(x + 101 < len(symbols)):
	        batch_symbols[i] = ",".join(symbols[x:x + 100])
	    else:
	        batch_symbols[i] = ",".join(symbols[x:len(symbols) + 1])
	        break
	    x = (i + 1) * 100 + 1

	return batch_symbols


def xl_get_stats():

    file = 'stock_score_data.xlsx'
    wb = xl.load_workbook(file)
    ws = wb['Data']
    symbols = get_symbols()
    stats = {}
    characteristics = ['marketcap', 'beta', 'week52high', 'week52low', \
    'shortInterest', 'dividendRate', 'dividendYield', 'latestEPS', 'sharesOutstanding', \
    'returnOnEquity', 'EBITDA', 'revenue', 'grossProfit', 'cash', 'debt', 'ttmEPS', \
    'revenuePerShare', 'revenuePerEmployee', 'peRatioHigh', 'peRatioLow', 'returnOnAssets', \
    'profitMargin', 'priceToSales', 'priceToBook', 'day200MovingAvg', 'day50MovingAvg', \
    'institutionPercent', 'year5ChangePercent', 'year2ChangePercent', 'year1ChangePercent', \
    'ytdChangePercent', 'month6ChangePercent', 'month3ChangePercent', 'month1ChangePercent', \
    'day5ChangePercent']

    for symbol in symbols:
	    for characteristic in characteristics:
	    	column = xl.get_column_letter(characteristics.index(characteristic) + 2)
	    	if ws[column + str(3)].value = characteristic:
			    for i in range(4,6000):
			        if ws[column + str(i)].value != None:
			            stats['symbol'][characteristic] = ws[column + str(i)].value
	return stats


def get_stats(iex_url_base = iex_url_base, manual = False):

	file = 'stock_score_data.xlsx'
    wb = xl.load_workbook(file)
    ws = wb['Data']
    symbols = xl_get_symbols()
    batch_symbols = set_batches(symbols)
    stats = {}
    characteristics = ['marketcap', 'beta', 'week52high', 'week52low', \
    'shortInterest', 'dividendRate', 'dividendYield', 'latestEPS', 'sharesOutstanding', \
    'returnOnEquity', 'EBITDA', 'revenue', 'grossProfit', 'cash', 'debt', 'ttmEPS', \
    'revenuePerShare', 'revenuePerEmployee', 'peRatioHigh', 'peRatioLow', 'returnOnAssets', \
    'profitMargin', 'priceToSales', 'priceToBook', 'day200MovingAvg', 'day50MovingAvg', \
    'institutionPercent', 'year5ChangePercent', 'year2ChangePercent', 'year1ChangePercent', \
    'ytdChangePercent', 'month6ChangePercent', 'month3ChangePercent', 'month1ChangePercent', \
    'day5ChangePercent']

    last_updated = ws['B1'].value
    time_now = datetime.datetime.now()
    time_diff = time_now - last_updated

    if (last_updated == None or time_diff > datetime.timedelta(days=5) or manual):

    	for i in batch_symbols:
	        batch_url = iex_url_base + "stock/market/batch?symbols=" + batch_symbols[i] + "&types=stats"
	        result = requests.get(batch_url).json()
	        print(i)
	        for symbol in result:
	            if(result[symbol.upper()].get('stats')):

	                stats[symbol] = result[symbol]['stats']

	    for symbol in symbols:
	        for characteristic in characteristics:
	            column = xl.utils.cell.get_column_letter(characteristics.index(characteristic) + 2)
	            row = symbols.index(symbol) + 4
	            ws[column + str(row)].value = stats[symbol][characteristic]
	            print(column + str(row), "value set to " + str(stats[symbol][characteristic]))
	    wb.save(file)

    else:

    	stats = xl_get_stats()

    return stats


def get_stats():


def total_setup():

	""" Total setup returns symbols, stock_scores, and batch_symbols.
	"""

	symbols = get_symbols()
	stock_scores = init_stock_scores(symbols)
	batch_symbols = set_batches(symbols)

	return symbols, stock_scores, batch_symbols


def soup_it(url):

	"""Returns html from specified url using Beautiful Soup.

	Must further strip to meaningfully use the returned html result.
	"""
	page = requests.get(url).text.encode("utf-8").decode('ascii', 'ignore')
	soup = BeautifulSoup(page, 'html.parser')

	return soup



def return_top(dict, x = None):

	if x == None:
		x = len(dict)
	sorted_array = sorted(dict.items(), key=lambda x: x[1], reverse = True)

	return sorted_array[0:x]


